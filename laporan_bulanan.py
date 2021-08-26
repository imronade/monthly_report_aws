# Made by 🤍
# Imron Ade
# August, 25 2021

# import modules
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import date, timedelta
import datetime
import os
from dateutil.relativedelta import relativedelta
import glob
from dateutil import tz
import sys
import shutil

# Setting up date
print("Checking folder..")
currentDate = date.today()
currentDate = currentDate.strftime("%Y%m%d")

CHECK_FOLDER = os.path.isdir("output/laporan_bulanan/"+currentDate)
# If folder doesn't exist, then create it.
if not CHECK_FOLDER:
    os.makedirs("output/laporan_bulanan/"+currentDate)
    print("created folder : ", "output/laporan_bulanan/"+currentDate)
else:
    print("output/laporan_bulanan/"+currentDate, "folder already exists.")

# Make temporary folder
CHECK_Temp_Folder = os.path.isdir("output/laporan_bulanan/temp")
# If folder doesn't exist, then create it.
if not CHECK_Temp_Folder:
    os.makedirs("output/laporan_bulanan/temp")
    print("created folder : ", "output/laporan_bulanan/temp")
else:
    print("output/laporan_bulanan/temp", "folder already exists.")

# print list from get_label
def print_list(listdata):
    for i in listdata:
        print(listdata.index(i),". ",i,sep="")
    # features select everything (under develop)
    print(listdata.index(i)+1,". ","Semuanya",sep="")

# Get list data from menu except all function
# def get_main_menu_choice(listdata):
#     while True:    
#         try:
#             number = int(input('Aku memilih: '))
#             if 0 <= number < len(listdata):
#                 return number
#             elif number == len(listdata):
#                 print("Menu Dalam Pengembangan")
#                 pass
#             elif number > len(listdata):
#                 print("Input Tidak Ada Dalam List Menu")
#                 pass
#         except (ValueError, TypeError):
#             print("Input Tidak Diketahui")
#             print("Silahkan Coba Lagi")
#             pass

# get metadata from choosen file
def metadata_waktu(index = None):
    iFile = pd.read_csv(os.path.join(path_data, csvFilenamesList[index]), sep=',')
    iFile['Tanggal'] = pd.to_datetime(iFile['Tanggal'])
    iFile['Tanggal'] =  iFile['Tanggal'].dt.tz_localize("GMT").dt.tz_convert('Asia/Jakarta')
    meta_bulan = iFile['Tanggal'].dt.month[0]
    meta_tahun = iFile['Tanggal'].dt.year[0]
    
    bulan_huruf = {1: 'JANUARI', 2:'FEBRUARI', 3:'MARET', 4:'APRIL',\
                5:'MEI', 6:'JUNI', 7:'JULI', 8:'AGUSTUS', 9:'SEPTEMBER',\
                10:'OKTOBER', 11:'NOVEMBER', 12:'DESEMBER'}

    # get periode
    
    if (meta_tahun % 4) == 0:
       if (meta_tahun % 100) == 0:
           if (meta_tahun % 400) == 0:
               kabisat = True
           else:
               kabisat = False
       else:
           kabisat = True
    else:
        kabisat = False
    
    # get name file
    name_file = csvFilenamesList[index]
    name_file = name_file.split("_",-1)[2] 
    tipe_alat = name_file.split(" ",1)[0]
    lokasi = name_file.split(" ",1)[1]

    meta_data =[meta_bulan, meta_tahun, kabisat, tipe_alat, lokasi]

    return meta_data, bulan_huruf

# open csv files
path = os.getcwd()
print("Folder yang tersedia..")
print(os.listdir(path+"\\output\\merged"))
alamat_data = int(input("Masukkan nama folder pada folder merged (20210419):  "))
path_data = path+"\\output\\merged\\"+str(alamat_data)+"\\"
path_simpan = path+"\\output\\laporan_bulanan\\"+currentDate+"\\"

# Listing files in folder
print("Listing files in folder..")
csvFilenamesList = os.listdir(path_data)
print("Done")
print("Pilih file yang akan dibuat untuk laporan bulanan : ")
print_list(csvFilenamesList)
ifile = int(input('Aku memilih: '))
max_files = len(csvFilenamesList)

# setting up monthly report format 
def open_format(bulan=None, tahun=None,kabisat=None):
    hari31 = [1,3,5,7,8,10,12]
    hari30 = [4,6,9,11]
    wb = load_workbook('master/master_laporan_bulanan.xlsx')
    sheets = wb.sheetnames
    start_date = datetime.date(tahun, bulan, 1)
    
    if bulan in hari31 :
        end_date = datetime.date(tahun,bulan,31)
        dates_31 = [ start_date + datetime.timedelta(n) for n in range(int ((end_date - start_date).days)+1)]
        date_strings = [d.strftime('%d/%m/%Y') for d in dates_31]

        for s in sheets:
            if s != '31_hari':
                sheet_name = wb.get_sheet_by_name(s)
                wb.remove_sheet(sheet_name)
            else:
                continue
        sheet = wb["31_hari"]
        col = "A"
        row = 12
        for i in range(0,31):
            y = str(row)
            cell = col + y
            sheet[cell] = str(date_strings[i])
            row += 1    
        sheet["B11"]= "00:00"
        sheet["B11"].alignment = Alignment(horizontal='right')
        wb.save("output/laporan_bulanan/temp/31_hari.xlsx")
    elif bulan in hari30 :
        end_date = datetime.date(tahun,bulan,30)
        dates_30 = [ start_date + datetime.timedelta(n) for n in range(int ((end_date - start_date).days)+1)]
        date_strings = [d.strftime('%d/%m/%Y') for d in dates_30]

        for s in sheets:
            if s != '30_hari':
                sheet_name = wb.get_sheet_by_name(s)
                wb.remove_sheet(sheet_name)
            else:
                continue
        sheet = wb["30_hari"]
        col = "A"
        row = 12
        for i in range(0,30):
            y = str(row)
            cell = col + y
            sheet[cell] = str(date_strings[i])
            row += 1    
        sheet["B11"]= "00:00"
        sheet["B11"].alignment = Alignment(horizontal='right')
        wb.save("output/laporan_bulanan/temp/30_hari.xlsx")
    elif bulan == 2 and kabisat == True :
        end_date = datetime.date(tahun,bulan,29)
        dates_29 = [ start_date + datetime.timedelta(n) for n in range(int ((end_date - start_date).days)+1)]
        date_strings = [d.strftime('%d/%m/%Y') for d in dates_29]

        for s in sheets:
            if s != '29_hari':
                sheet_name = wb.get_sheet_by_name(s)
                wb.remove_sheet(sheet_name)
            else:
                continue
        sheet = wb["29_hari"]
        col = "A"
        row = 12
        for i in range(0,29):
            y = str(row)
            cell = col + y
            sheet[cell] = str(date_strings[i])
            row += 1    
        sheet["B11"]= "00:00"
        sheet["B11"].alignment = Alignment(horizontal='right')
        wb.save("output/laporan_bulanan/temp/29_hari.xlsx")
    else:
        end_date = datetime.date(tahun,bulan,28)
        dates_28 = [ start_date + datetime.timedelta(n) for n in range(int ((end_date - start_date).days)+1)]
        date_strings = [d.strftime('%d/%m/%Y') for d in dates_28]

        for s in sheets:
            if s != '28_hari':
                sheet_name = wb.get_sheet_by_name(s)
                wb.remove_sheet(sheet_name)
            else:
                continue
        sheet = wb["28_hari"]
        col = "A"
        row = 12
        for i in range(0,28):
            y = str(row)
            cell = col + y
            sheet[cell] = str(date_strings[i])
            row += 1    
        sheet["B11"]= "00:00"
        sheet["B11"].alignment = Alignment(horizontal='right')
        wb.save("output/laporan_bulanan/temp/28_hari.xlsx")

varColomn = {'rr':'rr', 'ws':'ws_avg', 'wd':'wd_avg', 'ta':'tt_air_avg',
            'rh':'rh_avg', 'pp':'pp_air', 'sr':'sr_avg'}
varStat = {'rr':'last','ws_avg':'last', 'wd_avg':'last', 'tt_air_avg':'mean',
            'rh_avg':'mean', 'pp_air':'mean', 'sr_avg':'mean'}
varParam = {'rr':': CURAH HUJAN', 'ws':': KECEPATAN ANGIN', 'wd':'ARAH ANGIN', 'ta':'TEMPERATUR UDARA',
            'rh':': KELEMBAPAN', 'pp':': TEKANAN UDARA', 'sr':': PENYINARAN MATAHARI'}
varUnit = {'rr':': MILIMETER (mm)', 'ws':': m/s', 'wd':': DERAJAT', 'ta':': DERAJAT CELCIUS',
            'rh':': PERSEN', 'pp':': MILIBAR (mbar)', 'sr':': WATT/JAM'}

varFiles = {1: '31_hari.xlsx', 2:'28_hari.xlsx', 3:'31_hari.xlsx',
            4:'30_hari.xlsx', 5:'31_hari.xlsx', 6:'30_hari.xlsx',
            7:'31_hari.xlsx', 8:'31_hari.xlsx', 9:'30_hari.xlsx',
            10:'31_hari.xlsx', 11:'30_hari.xlsx', 12:'31_hari.xlsx'}

# alternative
# def key_mapper(obj, key):
#     if key == 'rr':
#       return obj.last()
#     if key == 'pp':
#       return obj.mean()

# myObj = MyClass()
# myObj = dataku.resample('1H', on='Tanggal')

def selectTemp(bulan=None, statusKabisat = None):
    if bulan in [1,3,5,7,8,10,12]:
        tempFile = '31_hari.xlsx'
    elif bulan in [4,6,9,11]:
        tempFile = '30_hari.xlsx'
    elif bulan == 2 and statusKabisat == True:
        tempFile = '29_hari.xlsx'
    else:
        tempFile = '28_hari.xlsx'
    return tempFile

def data_jam(data):
    dataku = pd.read_csv(data, sep=',')
    # change utc to local timezone
    dataku['Tanggal'] = pd.to_datetime(dataku['Tanggal'])
    dataku['Tanggal'] =  dataku['Tanggal'].dt.tz_localize("GMT").dt.tz_convert('Asia/Jakarta')

    # select varibales based on device typed
    if meta_data[3] == 'ARG':
        # delete unnecessary columns
        dataku = dataku.filter(['Tanggal', 'rr'])
        # processing to calculate hourly data
        dataku2 = dataku.resample('1H', on='Tanggal').last()
        dataku2 = dataku2[dataku2['Tanggal'].notna()]
        dataku2['year'] = [ts.year for ts in dataku2.index]
        dataku2['month'] = [ts.month for ts in dataku2.index]
        dataku2['day'] = [ts.day  for ts in dataku2.index]
        dataku2['hour'] = [ts.hour for ts in dataku2.index]
        data_rr = dataku2[dataku2['month'] == meta_data[0]]

        # make monthly report for rr
        path_temp = "output/laporan_bulanan/temp/"
        tempFile = selectTemp(bulan = meta_data[0], statusKabisat = meta_data[2])
        data_temp = os.path.join(path_temp, tempFile)
        wb = load_workbook(data_temp)
        
        # change sheet name
        ws = wb.active
        ws.title = "rr"
        
        # koordinat cell
        jam = list(range(0,24))
        kolom_jam = list(range(2,26))
        dict_waktu = dict(zip(jam, kolom_jam))
        if ws.cell(row=40, column=1).value == 'RATA-RATA':
            row_tanggal = list(range(12,40))
            tanggal = list(range(1,29))
        elif ws.cell(row=41, column=1).value == 'RATA-RATA':
            row_tanggal= list(range(12,41))
            tanggal = list(range(1,30))
        elif ws.cell(row=42, column=1).value == 'RATA-RATA':
            row_tanggal = list(range(12,42))
            tanggal = list(range(1,31))
        elif ws.cell(row=43, column=1).value == 'RATA-RATA':
            row_tanggal = list(range(12,43))
            tanggal = list(range(1,32))
        dict_date = dict(zip(tanggal, row_tanggal))
       
        # fill the value
        for t in range(0,len(data_rr)):
           tanggal_baris = dict_date[data_rr['day'][t]]
           waktu_kolom = dict_waktu[data_rr['hour'][t]]
           ws.cell(row=tanggal_baris, column=waktu_kolom).value = data_rr['rr'][t]
        
        # fill the metadata
        ws.cell(row= 4, column=3).value = ': CURAH HUJAN'
        ws.cell(row= 5, column=3).value = ': MILIMETER (mm)'
        ws.cell(row= 6, column=3).value = ': '+meta_data[3]
        ws.cell(row= 7, column=3).value = ': '+bulan_huruf[meta_data[0]]+' '+str(meta_data[1])
        ws.cell(row= 8, column=3).value = ': '+meta_data[4]
            
        # export the output
        outputName = [str(meta_data[3]),str(meta_data[4]), str(meta_data[1]), bulan_huruf[meta_data[0]]]
        outputName = '_'.join(outputName)
        wb.save(path_simpan+'rr_'+outputName+'.xlsx')         
        
    elif meta_data[3] == 'AWS' or meta_data[3] == 'AAWS':
        for key, value in varColomn.items():
            # processing to calculate hourly data
            dataku2 = dataku.resample('1H', on='Tanggal').agg(varStat)
            dataku2['Tanggal'] = dataku2.index
            dataku2 = dataku2[dataku2['Tanggal'].notna()]
            dataku2['year'] = [ts.year for ts in dataku2.index]
            dataku2['month'] = [ts.month for ts in dataku2.index]
            dataku2['day'] = [ts.day  for ts in dataku2.index]
            dataku2['hour'] = [ts.hour for ts in dataku2.index]
            dataku3 = dataku2[dataku2['month'] == meta_data[0]]
            
            # make monthly report for rr
            path_temp = "output/laporan_bulanan/temp/"
            tempFile = selectTemp(bulan = meta_data[0], statusKabisat = meta_data[2])
            data_temp = os.path.join(path_temp, tempFile)
            wb = load_workbook(data_temp)
            
            # change sheet name
            ws = wb.active
            ws.title = key
            
             # koordinat cell
            jam = list(range(0,24))
            kolom_jam = list(range(2,26))
            dict_waktu = dict(zip(jam, kolom_jam))
            if ws.cell(row=40, column=1).value == 'RATA-RATA':
                row_tanggal = list(range(12,40))
                tanggal = list(range(1,29))
            elif ws.cell(row=41, column=1).value == 'RATA-RATA':
                row_tanggal= list(range(12,41))
                tanggal = list(range(1,30))
            elif ws.cell(row=42, column=1).value == 'RATA-RATA':
                row_tanggal = list(range(12,42))
                tanggal = list(range(1,31))
            elif ws.cell(row=43, column=1).value == 'RATA-RATA':
                row_tanggal = list(range(12,43))
                tanggal = list(range(1,32))
            dict_date = dict(zip(tanggal, row_tanggal))
           
            # fill the value
            for t in range(0,len(dataku3)):
               tanggal_baris = dict_date[dataku3['day'][t]]
               waktu_kolom = dict_waktu[dataku3['hour'][t]]
               ws.cell(row=tanggal_baris, column=waktu_kolom).value = dataku3[value][t]
            
            # fill the metadata
            ws.cell(row= 4, column=3).value = varParam[key]
            ws.cell(row= 5, column=3).value = varUnit[key]
            ws.cell(row= 6, column=3).value = ': '+meta_data[3]
            ws.cell(row= 7, column=3).value = ': '+bulan_huruf[meta_data[0]]+' '+str(meta_data[1])
            ws.cell(row= 8, column=3).value = ': '+meta_data[4]
                
            # export the output
            outputName = [str(meta_data[3]),str(meta_data[4]), str(meta_data[1]), bulan_huruf[meta_data[0]]]
            outputName = '_'.join(outputName)
            wb.save(path_simpan+str(key)+'_'+outputName+'.xlsx')         
                   
    return

# processing data
if ifile < max_files:
    meta_data, bulan_huruf = metadata_waktu(ifile)
    open_format(bulan = meta_data[0], tahun= meta_data[1], kabisat=meta_data[2])
    print("Monthly report format has been created")
    data = os.path.join(path_data, csvFilenamesList[ifile])
    print('Processing to make monthly report..')
    data_jam(data)
    print('Process Done')       
elif ifile == max_files:
    for i in range(0,max_files):
        meta_data, bulan_huruf = metadata_waktu(i)
        open_format(bulan = meta_data[0], tahun= meta_data[1], kabisat=meta_data[2])
        print("Monthly report format has been created")
        data = os.path.join(path_data, csvFilenamesList[i])
        print('Processing to make monthly report..')
        data_jam(data)
        print(str(meta_data[3])+' '+str(meta_data[4])+' '+bulan_huruf[meta_data[0]]+' Report Created')
        print('Process Done')

print('Finish to make monthly report')
shutil.rmtree('output/laporan_bulanan/temp')
